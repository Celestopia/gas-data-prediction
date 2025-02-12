import numpy as np
import matplotlib
import pandas as pd
import warnings
import torch
import os
import sys
import time
import random
sys.path.append(os.path.join(os.path.dirname(os.path.abspath('__file__')), '..')) # Modify the working path so that this.ipynb file can import other modules like in the root directory
matplotlib.use('Agg') # Set the backend to disable figure display window
#warnings.filterwarnings("ignore", category=UserWarning) # To filter the warning of disabling plt.show()
t0=time.time() # Start the timer
print("\n##############======== main.py Starting ========##############\n")

# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# --- Parse command line arguments -----------------------------------------------------------------------------------------------------------------------------------------------------------
# --------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------



import argparse
parser = argparse.ArgumentParser(description='main.py')

# --- Dataset settings --------------------------------------------------------------
parser.add_argument('--data_path', type=str, default='E:\\科创优才\\实验数据\\天然气锅炉数据1.xlsx', help='data_path')
parser.add_argument('--sheet_name', type=str, default='1-2月份数据', help='sheet_name')
parser.add_argument('--input_var_names', type=str, default=
                                                    "主蒸汽流量计算值 \
                                                    锅炉天然气进气流量 \
                                                    锅炉天然气进气温度 \
                                                    锅炉天然气进气压力 \
                                                    鼓风机出口温度 \
                                                    鼓风机出口压力 \
                                                    鼓风机变频器输出反馈 \
                                                    鼓风机变频器电流反馈 \
                                                    冷凝器出口烟气调节阀反馈 \
                                                    SWY大气压 \
                                                    SWY天气温度 \
                                                    SWY空气湿度 \
                                                    SWY湿球温度 \
                                                    主蒸汽温度（蒸汽集箱出口温度） \
                                                    主蒸汽压力（蒸汽集箱出口压力）",
                                                help='The names of input variables.')
parser.add_argument('--output_var_names', type=str, default=
                                                    "烟气含氧量（CEMS） \
                                                    NOX标干浓度 \
                                                    一氧化碳",
                                                help='The names of output variables.')
parser.add_argument('--input_len', type=int, default=1, help='The temporal length of input sequence')
parser.add_argument('--output_len', type=int, default=1, help='The temporal length of output sequence')
parser.add_argument('--overlap', type=int, default=1, help='''The overlap between input and output sequence.\n
                            Usually set to 0 or 1.\n
                            0 means no overlap between input and output sequence;\n
                            1 means the last timestep of input sequence is the first timestep of output sequence.
                            ''')

# --- Model settings ----------------------------------------------------
parser.add_argument('--model_name', type=str, default='SVR', help='The model used for prediction.')


# SVR settings
parser.add_argument('--C', type=float, default=1.0, help='The parameter C of SVR.')
parser.add_argument('--epsilon', type=float, default=0.1, help='The parameter epsilon of Linear SVR.')
parser.add_argument('--nu', type=float, default=0.5, help='The parameter nu of NuSVR.')
parser.add_argument('--kernel', type=str, default='rbf', help='The kernel function of SVR. options: ["rbf", "linear", "poly", "sigmoid"].')
parser.add_argument('--degree', type=int, default=3, help='Degree of the polynomial kernel function ("poly"). Must be non-negative. Ignored by all other kernels.')

# LR settings
parser.add_argument('--l2_ratio', type=float, default=0.1, help='The scale parameter of L2 regularization.')
parser.add_argument('--l1_ratio', type=float, default=0.1, help='The scale parameter of L1 regularization.')

# Ensemble settings
parser.add_argument('--n_estimators', type=int, default=100, help='The number of trees in the forest.')
parser.add_argument('--max_depth', type=int, default=5, help='The maximum depth of the tree.')
parser.add_argument('--min_samples_split', type=int, default=2, help='The minimum number of samples required to split an internal node.')
parser.add_argument('--min_samples_leaf', type=int, default=1, help='The minimum number of samples required to be at a leaf node.')
parser.add_argument('--max_features', type=str, default='sqrt', help='The number of features to consider when looking for the best split. Options: ["auto", "sqrt", "log2"].')

# --- Training settings ---------------------------------------------

# General settings
parser.add_argument('--seed', type=int, default=20241215, help='The random seed.')
parser.add_argument('--device', type=str, default='cpu', help='device to train the neural network model. Options: ["cuda", "cpu"].')


# SVR settings
parser.add_argument('--tol', type=float, default=1e-3, help='The tolerance of the stopping criterion.')
parser.add_argument('--max_iter', type=int, default=-1, help='The maximum number of iterations. -1 means no limit.')

# NN settings
parser.add_argument('--optimizer_name', type=str, default='Adam', help='The optimizer used for training. Options: ["Adam", "SGD", "Adagrad", "Adadelta", "RMSprop"].')
parser.add_argument('--learning_rate', type=float, default=0.001, help='learning rate')
parser.add_argument('--batch_size', type=int, default=32, help='batch size')
parser.add_argument('--num_epochs', type=int, default=100, help='maximum number of epochs (may stop earlier due to early stopping settings)')

# Ensemble settings
parser.add_argument('--ensemble_learning_rate', type=float, default=0.1, help='The learning rate of the ensemble model. Used for gradient boosting.')

# --- Saving Settings ----------------------------------------------------
parser.add_argument('--save_record', type=int, default=1, help='Whether to save the record files. 1 for saving, 0 for not saving.')
parser.add_argument('--record_directory', type=str, default='./_results/', help='The directory to save the training record. The results of each run will be saved in a sub-directory under this directory.')
parser.add_argument('--figsize', type=str, default="18 12", help='The size of the figure.')

parser.add_argument('--save_log', type=int, default=1, help='Whether to save the training log. 1 for saving, 0 for not saving.')
parser.add_argument('--log_path', type=str, default='./_results/_log.xlsx', help='The path of the log file (must be .xlsx).')
parser.add_argument('--metadata', type=str, default='', help='Additional information to be saved within the logfile')

# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# --- Data Preparation ------------------------------------------------------------------------------------------------------------------------------------------------------------------
# ---------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------


args=parser.parse_args()

# Set the random seeds
random.seed(args.seed) # Seed for random
np.random.seed(args.seed) # Seed for numpy
torch.manual_seed(args.seed) # Seed for PyTorch of CPU
torch.cuda.manual_seed(args.seed) # Seed for PyTorch CUDA of GPU
torch.cuda.manual_seed_all(args.seed) # Seed for PyTorch CUDA of all GPUs
torch.backends.cudnn.deterministic = True # Deterministic mode for cuDNN

input_var_names=args.input_var_names.split()
output_var_names=args.output_var_names.split()

print("----------------Model Setting Information----------------")
print("Model name: ", args.model_name)
print("Input variables:\n", input_var_names)
print("Output variables:\n", output_var_names)
print("Input length: ", args.input_len)
print("Output length: ", args.output_len)
print("Input channels: ", len(input_var_names))
print("Output channels: ", len(output_var_names))
print("Overlap: ", args.overlap)
print("max_iter: ", args.max_iter)
print("Record directory: ", args.record_directory)
print("Log path: ", args.log_path)
print("Seed: ", args.seed)
print("---------------------------------------------------------")


# --- Load data -----------------------------------------------------------------------
from data_provider.data_reading import load_data

DATA, \
    (var_names, var_units), \
    (input_var_units, output_var_units), \
    (input_var_indices, output_var_indices)=load_data(args.data_path,
                                                        args.sheet_name,
                                                        input_var_names,
                                                        output_var_names)


# --- Hyperparameters -----------------------------------------------------------------------------------------
input_len = args.input_len
output_len = args.output_len
overlap = args.overlap # 1; The last time step of the input sequence is the first time step of the output sequence.
input_channels = len(input_var_names)
output_channels = len(output_var_names)


# --- Data preprocessing -----------------------------------------------------------------------------------------
from data_provider.data_preprocessing import GasData

DATASET=GasData(DATA,
                input_len=input_len,
                output_len=output_len,
                overlap=overlap,
                input_indices=input_var_indices,
                output_indices=output_var_indices,
                var_names=var_names,
                var_units=var_units)
transformed_data, \
    (var_mean, var_std_dev), \
    (input_var_mean, input_var_std_dev), \
    (output_var_mean, output_var_std_dev) \
        = DATASET.standardize()
DATASET.train_test_split()
(X_train_grouped, Y_train_grouped), (X_test_grouped, Y_test_grouped) = DATASET.time_series_slice()
(X_train, Y_train), (X_test, Y_test) = DATASET.build_train_test_set()



# -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# --- Model building and training ---------------------------------------------------------------------------------------------------------------------------------------------------------
# -----------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# --- Determine the model type ------------------------------------------------------------------------------

model_name=args.model_name
model_type=None
if model_name in ['MLP', 'CNN', 'LSTM', 'RNN', 'TCN', 'GRU']:
    model_type='NN'
elif model_name in ['SVR', 'LinearSVR', 'NuSVR']:
    model_type='SVR'
elif model_name in ['LR', 'Lasso', 'Ridge', 'ElasticNet']:
    model_type='LR'
elif model_name in ['RandomForest', 'ExtraTrees', 'AdaBoost', 'Bagging', 'GradientBoosting']:
    model_type='Ensemble'
elif model_name in ['GPR']:
    model_type='GPR'
else:
    raise ValueError(f"{args.model_name} is an invalid model name.")


# --- Train model based on the model type ------------------------------------------------------------------------------

shape_params = {
    'input_len': input_len,
    'output_len': output_len,
    'input_channels': input_channels,
    'output_channels': output_channels,
}

if model_type=='NN':
    import torch.optim as optim
    import torch.nn as nn
    from models.RNN import RNN, LSTM, GRU
    from models.CNN import CNN, TCN
    from models.MLP import MLP
    from models.transformer import Transformer, iTransformer, PatchTST, Reformer, Informer
    from models.Linear import LLinear, DLinear, NLinear
    model_dict={
        'MLP': MLP,
        'CNN': CNN,
        'TCN': TCN,
        'RNN': RNN,
        'LSTM': LSTM,
        'GRU': GRU,
    }
    optimizer_dict={
        'Adam': optim.Adam,
        'SGD': optim.SGD,
        'Adagrad': optim.Adagrad,
        'Adadelta': optim.Adadelta,
        'RMSprop': optim.RMSprop
    }
    model=model_dict[model_name](**shape_params).to(args.device)
    optimizer = optimizer_dict[args.optimizer_name](model.parameters(), lr=args.learning_rate)
    print('Number of NN model parameters: ', sum(p.numel() for p in model.parameters()))

    from utils.fit_history import FitHistory
    from utils.train import train
    from data_provider.data_preprocessing import get_XY_loaders
    train_loader, val_loader, test_loader = get_XY_loaders(X_train, Y_train,
                                                            train_ratio=0.8,
                                                            val_ratio=0.2,
                                                            test_ratio=0.0,
                                                            batch_size=args.batch_size)
    # train the model
    history=FitHistory()
    history.update(
                *train(model, train_loader, val_loader, optimizer,
                    loss_func=nn.MSELoss(),
                    metric_func=nn.L1Loss(),
                    num_epochs=args.num_epochs,
                    device=args.device,
                    verbose=1)
                )
    history.summary()


elif model_type=='SVR':
    from models.SVR import SVR, LinearSVR, NuSVR
    model_dict={
    'SVR': SVR(**shape_params, C=args.C, epsilon=args.epsilon, kernel=args.kernel, degree=args.degree, tol=args.tol, max_iter=args.max_iter),
    'LinearSVR': LinearSVR(**shape_params, C=args.C, tol=args.tol, max_iter=args.max_iter),
    'NuSVR': NuSVR(**shape_params, C=args.C, nu=args.nu, kernel=args.kernel, degree=args.degree, tol=args.tol, max_iter=args.max_iter),
    }
    model=model_dict[model_name]
    print("Fitting {} model...".format(model_name))
    model.fit(X_train,Y_train)


elif model_type=='LR':
    from models.LR import LinearRegression, Ridge, Lasso, ElasticNet
    model_dict={
        "LR": LinearRegression(**shape_params),
        "Ridge": Ridge(**shape_params,alpha=args.l2_ratio),
        "Lasso": Lasso(**shape_params,alpha=args.l1_ratio),
        "ElasticNet": ElasticNet(**shape_params,
                                    alpha=args.l1_ratio+args.l2_ratio,
                                    l1_ratio=args.l1_ratio/(args.l1_ratio+args.l2_ratio) if args.l1_ratio+args.l2_ratio!=0 else 0
                                    ), # See https://scikit-learn.org/stable/modules/generated/sklearn.linear_model.ElasticNet.html for parameter definition.
    }
    model=model_dict[model_name]
    print("Fitting {} model...".format(model_name))
    model.fit(X_train,Y_train)


elif model_type=='Ensemble':
    from models.Ensemble import RandomForestRegressor, ExtraTreesRegressor, AdaBoostRegressor, BaggingRegressor, GradientBoostingRegressor
    tree_params={
        "n_estimators": args.n_estimators,
        "max_depth": args.max_depth,
        "min_samples_split": args.min_samples_split,
        "min_samples_leaf": args.min_samples_leaf,
        "max_features": args.max_features,
    }
    model_dict={
        "RandomForest": RandomForestRegressor(**shape_params, **tree_params),
        "ExtraTrees": ExtraTreesRegressor(**shape_params, **tree_params),
        "AdaBoost": AdaBoostRegressor(**shape_params, n_estimators=args.n_estimators, learning_rate=args.ensemble_learning_rate),
        "Bagging": BaggingRegressor(**shape_params, n_estimators=args.n_estimators, max_samples=1.0, max_features=1.0),
        "GradientBoosting": GradientBoostingRegressor(**shape_params, **tree_params, learning_rate=args.ensemble_learning_rate),
    }
    model=model_dict[model_name]
    print("Fitting {} model...".format(model_name))
    model.fit(X_train,Y_train)


elif model_type=='GPR':
    from models.GPR import GaussianProcessRegressor
    model_dict={
        "GPR": GaussianProcessRegressor(**shape_params),
    }
    model=model_dict[model_name]
    print("Fitting {} model...".format(model_name))
    model.fit(X_train,Y_train)


#-------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------
# --- Model evaluation, visualization, and figure saving -----------------------------------------------------------------------------------------------------------------------------
# ------------------------------------------------------------------------------------------------------------------------------------------------------------------------------------

# --- Set up the directory to save the results and logs ----------
record_directory=args.record_directory
time_string=str(int(time.time()))
subdirectory='experiment_{}_{}'.format(
                                time_string,
                                model_name) # Use current timestamp to name subdirectory

# --- Model evaluation -------------------------------
from utils.model_test import ModelTest

Exp=ModelTest(model, DATASET, device=args.device) # Experiment object initialization

with_Tensor=True if model_type=='NN' else False
Y_pred, Y_true = Exp.get_pred_true_pairs(X_test_grouped[0], Y_test_grouped[0], with_Tensor=with_Tensor) # visualize on the first sample of test set
prediction_info=Exp.get_prediction_info(X_test_grouped,Y_test_grouped,with_Tensor=with_Tensor)


# --- Save model information ------------------------------------------------------------------------------------

model_info={
    "model_name": model_name,
    "input_len": input_len,
    "output_len": output_len,
    "input_channels": input_channels,
    "output_channels": output_channels,
    "overlap": overlap,
    "input variable names": input_var_names,
    "output variable names": output_var_names,
}
if model_type=='NN':
    model_info.update({
        "hyperparameters": "optimizer_name: {}, learning_rate: {}, batch_size: {}, num_epochs: {}".format(
                                args.optimizer_name,
                                args.learning_rate,
                                args.batch_size,
                                args.num_epochs),})
elif model_type=='SVR':
    model_info.update({
        "hyperparameters": "C: {}, epsilon: {}, nu: {}, kernel: {}, degree: {}, tol: {}, max_iter: {}".format(
                                args.C,
                                args.epsilon,
                                args.nu,
                                args.kernel,
                                args.degree,
                                args.tol,
                                args.max_iter),})
elif model_type=='LR':
    model_info.update({
        "hyperparameters": "l1_ratio: {}, l2_ratio: {}".format(
                                args.l1_ratio,
                                args.l2_ratio),})
elif model_type=='Ensemble':
    model_info.update({
        "hyperparameters": "n_estimators: {}, max_depth: {}, min_samples_split: {}, min_samples_leaf: {}, max_features: {}, ensemble_learning_rate: {}".format(
                                args.n_estimators,
                                args.max_depth,
                                args.min_samples_split,
                                args.min_samples_leaf,
                                args.max_features,
                                args.ensemble_learning_rate),})
elif model_type=='GPR':
    model_info.update({
        "hyperparameters": "default",})


if args.save_record:
    print('Saving Record...')
    if not os.path.exists(os.path.join(record_directory, subdirectory)):
        os.makedirs(os.path.join(record_directory, subdirectory))

    # Save the plot of predictions
    for plot_name, (p,r) in zip(["pred_std","pred","res_std","res"], [(0,0), (0,1), (1,0), (1,1)]):
        Exp.plot_all_predictions(Y_pred, Y_true,
                                plot_residual=p,
                                rescale=r,
                                figsize=tuple(map(int, (args.figsize).split())),
                                save_path='{}/{}/{}_{}.png'.format(
                                    record_directory,
                                    subdirectory,
                                    model_name,
                                    plot_name))

    # Save the model information into a .json file
    import json
    model_info_json_save_path = '{}/{}/{}_model_info.json'.format(
                                    record_directory,
                                    subdirectory,
                                    model_name)
    with open(model_info_json_save_path, 'w') as json_file:
        json.dump(model_info, json_file, indent=4) 
    print(f"Saved model information to {model_info_json_save_path}")

    # Save prediction information
    import pandas as pd
    prediction_info_save_path = '{}/{}/{}_prediction_info.csv'.format(
                                    record_directory,
                                    subdirectory,
                                    model_name)
    prediction_info.to_csv(prediction_info_save_path, index=True, header=True)
    print(f"Saved prediction information to {prediction_info_save_path}")


# Save results to a conclusive .xlsx file, where results of different trainings are integrated, and can be easily compared.
if args.save_log:
    print('Saving Log...')
    import openpyxl

    sheet_title=['model_name','RMSE','MAE','RMSE_standard','MAE_standard','input_len','output_len','overlap','input variables','output variables']
    sheet_title.extend(["hyperparameters"])
    '''
    if model_type=='NN':
        sheet_title.extend(['learning_rate','batch_size','num_epochs','optimizer_name'])
    elif model_type=='SVR':
        sheet_title.extend(['C','epsilon','nu','kernel','degree','tol'])
    elif model_type=='LR':
        sheet_title.extend(['l1_ratio','l2_ratio',])'''
    sheet_title.extend(['time_string, seed, metadata'])

    if not os.path.exists(args.log_path): # Examine whether the log file exists. If not, create a new one.
        log_path_df=pd.DataFrame(columns=sheet_title)
        log_path_df.to_excel(args.log_path, index=False, header=True, sheet_name=output_var_names[0])
        print("Created a new log file at {}".format(args.log_path))

    workbook = openpyxl.load_workbook(args.log_path)
    sheet_names=workbook.sheetnames
    
    for var_name in output_var_names: # Examine whether the sheet for the current output variable exists. If not, create a new one.
        if var_name not in sheet_names:
            new_sheet=workbook.create_sheet(title=var_name)
            new_sheet.append(sheet_title)
            print("Created a new sheet for `{}` in log file at {}".format(var_name, args.log_path))

    for var_name in output_var_names: # Iterate over all output variables and save results to the corresponding sheet in the record file.
        sheet=workbook[var_name]
        row=1
        var_pred_info=prediction_info.loc[var_name]
        sheet.append({
            "A": model_name,
            "B": var_pred_info['RMSE'],
            "C": var_pred_info['MAE'],
            "D": var_pred_info['RMSE_standard'],
            "E": var_pred_info['MAE_standard'],
            "F": input_len,
            "G": output_len,
            "H": overlap,
            "I": ', '.join(input_var_names),
            "J": ', '.join(output_var_names),
            "K": model_info['hyperparameters'],
            "L": time_string,
            "M": args.seed,
            "N": args.metadata,})

        workbook.save(args.log_path)
        print("Saved results for `{}` to log file at {}".format(var_name, args.log_path))


print("Successfully finished! Running time: {:.2f} seconds.".format(time.time()-t0))
print("##############========= main.py Ending =========##############\n")


