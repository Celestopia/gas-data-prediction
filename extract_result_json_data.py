r"""
Extract the experiment data stored in .json files under a directory, and summarize them into an Excel file.
"""
import os
import json
import pandas as pd
import openpyxl

def get_json_paths(directory, suffix='result.json'):
    """
    Recursively find all result json files with certain suffix under a directory.

    :param directory: str, directory to search for json files.
    :param suffix: str, suffix of the json files to search for.
    :return: list of str, json file paths.
    """
    json_paths = []
    for root, dirs, files in os.walk(directory):
        for file in files:
            if file.endswith(suffix):
                json_paths.append(os.path.join(root, file))
    return json_paths


def extract_data_from_json(json_paths):
    """
    Extract data from a list of.json file paths.

    :param json_paths: list of str, json file paths.
    :return: a dictionary of variable names and their corresponding data.
    
    Example:
    ```python
    result_dict = {
        "time_string": "1700000000",
        "model_name": "CNN",
        "dataset_info": {
            "input_len": 10,
            "output_len": 1,
            "input_channels": 15,
            "output_channels": 2,
            "overlap": 1,
            "input_var_names": [
                "input_var_0",
                "input_var_1",
                "input_var_2",
            ],
            "output_var_names": [
                "output_var_0",
                "output_var_1",
            ]
        },
        "prediction_info": [
            {
                "var_name": "output_var_0",
                "var_unit": "mg/m3",
                "RMSE": 0.5076594568899324,
                "RMSE_rescaled": 2.897754015687841,
                "MAE": 0.40525201363247476,
                "MAE_rescaled": 2.313205504066275
            },
            {
                "var_name": "output_var_1",
                "var_unit": "%",
                "RMSE": 0.842616671563331,
                "RMSE_rescaled": 0.48989939262410753,
                "MAE": 0.6297549376913725,
                "MAE_rescaled": 0.36614106021025716
            }
        ],
        "hyperparameters": {
            "param_0": 0.1,
            "param_1": 0.3
        },
        "seed": 12345678
    }
    ```
    """
    data = {} # dictionary of lists, where each list contains data of a variable
    for json_path in json_paths:
        print('\n'+json_path)
        with open(json_path, 'r', encoding='utf-8') as file:
            result_dict = json.load(file) # dict
            for var_info in result_dict["prediction_info"]: # Loop over all variables in the json file
                var_name = var_info['var_name']
                print(var_name)
                row = {
                    'model_name': str(result_dict['model_name']),
                    'var_name': str(var_info['var_name']),
                    'var_unit': str(var_info['var_unit']),
                    'RMSE': str(var_info['RMSE']),
                    'MAE': str(var_info['MAE']),
                    'RMSE_rescaled': str(var_info['RMSE_rescaled']),
                    'MAE_rescaled': str(var_info['MAE_rescaled']),
                    'dataset_info': str(result_dict['dataset_info']),
                    'hyperparameters': str(result_dict['hyperparameters']),
                    'time_string': str(result_dict['time_string']),
                    'seed': str(result_dict['seed']),
                    }
                try:
                    data[var_name].append(row)
                except KeyError: # If the variable is not a key of the data dictionary yet
                    data[var_name] = []
                    data[var_name].append(row)
    return data


def save_to_xlsx(data, xlsx_save_path):
    """
    Save data to an Excel file.

    :param data: list of dictionaries.
    :param xlsx_save_path: str, path to the Excel file to save data to.
    """
    var_names = list(data.keys())
    sheet_title = ['model_name', 'var_name', 'var_unit', 'RMSE', 'MAE', 'RMSE_rescaled', 'MAE_rescaled', 'dataset_info', 'hyperparameters', 'time_string','seed']
    
    # Examine whether the summary file exists. If not, create a new one.
    if not os.path.exists(xlsx_save_path):
        df=pd.DataFrame(columns=sheet_title)
        df.to_excel(xlsx_save_path, index=False, header=True, sheet_name=var_names[0])
        print("Created a summary file at {}".format(xlsx_save_path))
        print("Created a new sheet for `{}` in summary file at {}".format(var_names[0], xlsx_save_path))

    workbook = openpyxl.load_workbook(xlsx_save_path)
    
    # Examine whether the sheet for the current output variable exists. If not, create a new one.
    for var_name in var_names:
        if var_name not in workbook.sheetnames:
            new_sheet=workbook.create_sheet(title=var_name)
            new_sheet.append(sheet_title)
            print("Created a new sheet for `{}` in summary file at {}".format(var_name, xlsx_save_path))
    
    # Save data to the corresponding sheet in the summary file.
    for var_name in var_names:
        sheet=workbook[var_name]
        for var_data_i in data[var_name]: # Loop over different experiment data of a variable; var_data_i: dictionary.
            # If the data already exists in the sheet, skip it to avoid duplicates.
            if var_data_i['time_string'] in [sheet.cell(row=i, column=10).value for i in range(1, sheet.max_row + 1)]:
                continue
            sheet.append({
                "A": var_data_i["model_name"],
                "B": var_data_i['var_name'],
                "C": var_data_i['var_unit'],
                "D": var_data_i['RMSE'],
                "E": var_data_i['MAE'],
                "F": var_data_i['RMSE_rescaled'],
                "G": var_data_i['MAE_rescaled'],
                "H": var_data_i['dataset_info'],
                "I": var_data_i['hyperparameters'],
                "J": var_data_i['time_string'],
                "K": var_data_i['seed'],})
        print(f"{var_name} data saved to {xlsx_save_path}; sheet name: {var_name}")

    workbook.save(xlsx_save_path) # Save all changes
    print(f"Summary file saved to {xlsx_save_path}")

if __name__ == '__main__':
    directory = r"E:\PythonProjects\gas-data-prediction\results20250302"
    xlsx_save_path = '20250302_summary.xlsx'
    json_paths = get_json_paths(directory)
    data = extract_data_from_json(json_paths)
    save_to_xlsx(data, xlsx_save_path)
    print('Done.')
