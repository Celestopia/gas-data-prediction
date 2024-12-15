import numpy as np
import matplotlib.pyplot as plt
import matplotlib
matplotlib.use('Agg') # Set the backend to disable figure display window
import warnings
warnings.filterwarnings("ignore", category=UserWarning) # To filter the warning of disabling plt.show()
import pandas as pd
import torch
import torch.nn as nn
import torch.optim as optim
import os
import sys
import time
import random
import seaborn as sns
from matplotlib.font_manager import FontProperties
font1 = FontProperties(fname=r"C:\\Windows\\Fonts\\STZHONGS.ttf", size=14)
font2 = FontProperties(fname=r"C:\\Windows\\Fonts\\STZHONGS.ttf", size=12)
font3 = FontProperties(fname=r"C:\\Windows\\Fonts\\STZHONGS.ttf", size=10)
font4 = FontProperties(fname=r"C:\\Windows\\Fonts\\STZHONGS.ttf", size=7)
sns.set_style('whitegrid')
sns.set_palette("muted")
random.seed(20241130)
np.random.seed(20241130)
torch.manual_seed(3407) # Torch.manual_seed(3407) is all you need. Paper: http://arxiv.org/abs/2109.08203
# Modify the working path so that this.ipynb file can import other modules like in the root directory
current_dir = os.path.dirname(os.path.abspath('__file__'))
sys.path.append(os.path.join(current_dir, '..'))


# --------------------------------------------------------------------------
# Parse command line arguments
# Open Git Bash in the directory of this file, and run ./LR_run_script.sh
import argparse

parser = argparse.ArgumentParser(description='NN_exe')
parser.add_argument('--model_name', type=str, default='CNN', help='The model used for prediction.')
parser.add_argument('--data_path', type=str, default='E:\\科创优才\\实验数据\\天然气锅炉数据1.xlsx', help='data_path')
parser.add_argument('--sheet_name', type=str, default='稳定运行数据段', help='sheet_name')

parser.add_argument('--input_var_names', type=str, default=
                                                    "主蒸汽流量计算值 \
                                                    锅炉天然气进气流量 \
                                                    锅炉天然气进气温度 \
                                                    锅炉天然气进气压力 \
                                                    鼓风机出口温度 \
                                                    鼓风机出口压力 \
                                                    鼓风机变频器输出反馈 \
                                                    鼓风机变频器电流反馈 \
                                                    冷凝器出口烟气调节阀反馈 \
                                                    SWY大气压 \
                                                    SWY天气温度 \
                                                    SWY空气湿度 \
                                                    SWY湿球温度 \
                                                    主蒸汽温度（蒸汽集箱出口温度） \
                                                    主蒸汽压力（蒸汽集箱出口压力）",
                                                help='The names of input variables.')

parser.add_argument('--output_var_names', type=str, default=
                                                    "烟气含氧量（CEMS） \
                                                    NOX浓度 \
                                                    一氧化碳",
                                                help='The names of output variables.')


parser.add_argument('--input_len', type=int, default=1, help='The temporal length of input sequence')
parser.add_argument('--output_len', type=int, default=1, help='The temporal length of output sequence')
parser.add_argument('--overlap', type=int, default=1, help='''The overlap between input and output sequence.\n
                            Usually set to 0 or 1.\n
                            0 means no overlap between input and output sequence;\n
                            1 means the last timestep of input sequence is the first timestep of output sequence.
                            ''')

parser.add_argument('--optimizer_name', type=str, default='Adam', help='The optimizer used for training. Options: ["Adam", "SGD", "Adagrad", "Adadelta", "RMSprop"].')
parser.add_argument('--learning_rate', type=float, default=0.001, help='learning rate')
parser.add_argument('--batch_size', type=int, default=32, help='batch size')
parser.add_argument('--num_epochs', type=int, default=200, help='Maximum number of epochs (but early stopping may be triggered)')
parser.add_argument('--device', type=str, default='cpu', help='device to train the neural network model. Options: ["cuda", "cpu"].')


parser.add_argument('--figsize', type=str, default="18 12", help='The size of the figure.')
parser.add_argument('--save_directory', type=str, default='./results', help='The directory to save the results.')

parser.add_argument('--save_record', type=bool, default=True, help='Whether to save the training record.')
parser.add_argument('--record_path', type=str, default='./results/NN_record.xlsx', help='The path to save the training record.')





args=parser.parse_args()

input_var_names=args.input_var_names.split()
output_var_names=args.output_var_names.split()

print("----------------Model Setting Information----------------")
print("Model name: ", args.model_name)
print("Input variables:\n", input_var_names)
print("Output variables:\n", output_var_names)
print("Input length: ", args.input_len)
print("Output length: ", args.output_len)
print("Input channels: ", len(input_var_names))
print("Output channels: ", len(output_var_names))
print("Overlap: ", args.overlap)
print("Optimizer: ", args.optimizer_name)
print("Learning rate: ", args.learning_rate)
print("Batch size: ", args.batch_size)
print("Number of epochs: ", args.num_epochs)
print("Save directory: ", args.save_directory)
print("Record path: ", args.record_path)
print("---------------------------------------------------------")


# --------------------------------------------------------------------------
# Load data
from gas_data_prediction.data_reading import load_data

DATA, \
    (var_names, var_units), \
    (input_var_units, output_var_units), \
    (input_var_indices, output_var_indices)=load_data(args.data_path, # Default: 'E:\\科创优才\\实验数据\\天然气锅炉数据1.xlsx'
                                                        args.sheet_name, # Default: '稳定运行数据段'
                                                        input_var_names,
                                                        output_var_names)


# --------------------------------------------------------------------------------------------
# Hyperparameters

input_len = args.input_len
output_len = args.output_len
overlap=args.overlap # The last time step of the input sequence is the first time step of the output sequence.
input_channels = len(input_var_names)
output_channels = len(output_var_names)


# --------------------------------------------------------------------------------------------
# Data preprocessing

from gas_data_prediction.data_preprocessing import GasData

DATASET=GasData(DATA,
                input_len=input_len,
                output_len=output_len,
                overlap=overlap,
                input_indices=input_var_indices,
                output_indices=output_var_indices,
                var_names=var_names,
                var_units=var_units)

transformed_data, \
    (var_mean, var_std_dev), \
    (input_var_mean, input_var_std_dev), \
    (output_var_mean, output_var_std_dev) \
        = DATASET.standardize()

DATASET.train_test_split()

(X_train_grouped, Y_train_grouped), (X_test_grouped, Y_test_grouped) = DATASET.time_series_slice()
(X_train, Y_train), (X_test, Y_test) = DATASET.build_train_test_set()


from gas_data_prediction.data_preprocessing import get_XY_loaders
train_loader, val_loader, test_loader = get_XY_loaders(X_train, Y_train, batch_size=args.batch_size)


# --------------------------------------------------------------------------------------------
# Model building and training

from models.RNN import RNN, LSTM, GRU
from models.CNN import CNN, TCN
from models.MLP import MLP
from models.transformer import Transformer, iTransformer, PatchTST, Reformer, Informer
from models.Linear import LLinear, DLinear, NLinear



model_name=args.model_name
learning_rate=args.learning_rate
batch_size=args.batch_size
num_epochs=args.num_epochs
optimizer_name=args.optimizer_name
device=args.device


shared_params = {
    'input_len': input_len,
    'output_len': output_len,
    'input_channels': input_channels,
    'output_channels': output_channels,
}

model_dict={
    'MLP': MLP(**shared_params).to(device),
    'CNN': CNN(**shared_params).to(device),
    'TCN': TCN(**shared_params).to(device),
    'RNN': RNN(**shared_params).to(device),
    'LSTM': LSTM(**shared_params).to(device),
    'GRU': GRU(**shared_params).to(device),
}


optimizer_dict={
    'Adam': optim.Adam,
    'SGD': optim.SGD,
    'Adagrad': optim.Adagrad,
    'Adadelta': optim.Adadelta,
    'RMSprop': optim.RMSprop
}

model=model_dict[model_name]
optimizer = optimizer_dict[optimizer_name](model.parameters(), lr=learning_rate)
print('Model parameters: ', sum(p.numel() for p in model.parameters()))

from gas_data_prediction.utils import FitHistory, train
# train the model
history=FitHistory()
history.update(
            *train(model, train_loader, val_loader, optimizer,
                loss_func=nn.MSELoss(),
                metric_func=nn.L1Loss(),
                num_epochs=num_epochs,
                device=device,
                verbose=1)
            )
history.summary()
#print("Loss on test set: ", model.evaluate((X_test, Y_test),mode="numpy",device=device))



# --------------------------------------------------------------------------------------------
# Model evaluation, visualization, and figure saving

from gas_data_prediction.utils import ModelTest

figsize=tuple(map(int, (args.figsize).split()))
save_directory=args.save_directory
time_string=str(int(time.time()))
subdirectory='experiment_{}_{}'.format(
                                time_string,
                                model_name) # Use current timestamp to name subdirectory
if not os.path.exists(os.path.join(save_directory, subdirectory)):
    os.makedirs(os.path.join(save_directory, subdirectory))

Exp=ModelTest(model, DATASET, device=device) # Experiment object initialization
Y_pred, Y_true = Exp.get_pred_true_pairs(X_test_grouped[0], Y_test_grouped[0], with_Tensor=True) # visualize on the first sample of test set

for plot_name, (p,r) in zip(["pred_std","pred","res_std","res"], [(0,0), (0,1), (1,0), (1,1)]):
    Exp.plot_all_predictions(Y_pred, Y_true,
                            plot_residual=p,
                            rescale=r,
                            figsize=figsize,
                            save_path='{}/{}/{}_{}.png'.format(
                                save_directory,
                                subdirectory,
                                model_name,
                                plot_name))

prediction_info=Exp.get_prediction_info(X_test_grouped,Y_test_grouped,with_Tensor=True)


# ---------------------------------------------------------------------------------------
# Save model information
import json

model_info={
    "model_name": model_name,
    "input_len": input_len,
    "output_len": output_len,
    "input_channels": input_channels,
    "output_channels": output_channels,
    "overlap": overlap,
    "input variables": input_var_names,
    "output variables": output_var_names,
    "learning_rate": learning_rate,
    "batch_size": batch_size,
    "num_epochs": history.num_epochs,
    "optimizer_name": optimizer_name,
    "device": device,
}
model_info_json_save_path = '{}/{}/{}_model_info.json'.format(
                                save_directory,
                                subdirectory,
                                model_name)
with open(model_info_json_save_path, 'w') as json_file:
    json.dump(model_info, json_file, indent=4) 
print(f"Saved model information to {model_info_json_save_path}")



# Save training record

history_dict={
    'num_epochs': history.num_epochs,
    'epoch_time': history.epoch_time,
    'train_loss': history.train_loss,
    'val_loss': history.val_loss,
    'train_metric': history.train_metric,
    'val_metric': history.val_metric,
}

history_info_json_save_path = '{}/{}/{}_train_history.json'.format(
                                save_directory,
                                subdirectory,
                                model_name)
with open(history_info_json_save_path, 'w') as json_file:
    json.dump(history_dict, json_file, indent=4) 
print(f"Saved training record to {history_info_json_save_path}")


# Save the figure of training curve

history.plot(save_path='{}/{}/{}_train_history_plot.png'.format(
                                save_directory,
                                subdirectory,
                                model_name))

# Save prediction information
import pandas as pd

prediction_info_save_path = '{}/{}/{}_prediction_info.csv'.format(
                                save_directory,
                                subdirectory,
                                model_name)

prediction_info.to_csv(prediction_info_save_path, index=True, header=True)
print(f"Saved prediction information to {prediction_info_save_path}")


# Save results to a conclusive .xlsx file, where results of different trainings are integrated, and can be easily compared.
import openpyxl

sheet_title=['model_name',
             'RMSE',
             'MAE',
             'RMSE_standard',
             'MAE_standard',
             'input_len',
             'output_len',
             'overlap',
             'input variables',
             'output variables', 
             'learning_rate',
             'batch_size',
             'num_epochs',
             'optimizer_name',
             'time_string']
if args.save_record:
    if not os.path.exists(args.record_path):
        record_path_df=pd.DataFrame(columns=sheet_title)
        record_path_df.to_excel(args.record_path, index=False, header=True, sheet_name=output_var_names[0])
        print("Created a new record file at {}".format(args.record_path))

    workbook = openpyxl.load_workbook(args.record_path)
    sheet_names=workbook.sheetnames
    
    for var_name in output_var_names:
        if var_name not in sheet_names:
            new_sheet=workbook.create_sheet(title=var_name)
            new_sheet.append(sheet_title)
            print("Created a new sheet for `{}` in record file at {}".format(var_name, args.record_path))

    for var_name in output_var_names:
        sheet=workbook[var_name]
        row=1
        var_pred_info=prediction_info.loc[var_name]
        sheet.append({
            "A": model_name,
            "B": var_pred_info['RMSE'],
            "C": var_pred_info['MAE'],
            "D": var_pred_info['RMSE_standard'],
            "E": var_pred_info['MAE_standard'],
            "F": input_len,
            "G": output_len,
            "H": overlap,
            "I": ', '.join(input_var_names),
            "J": ', '.join(output_var_names),
            "K": learning_rate,
            "L": batch_size,
            "M": num_epochs,
            "N": optimizer_name,
            "O": time_string})

        workbook.save(args.record_path)
        print("Saved results for `{}` to record file at {}".format(var_name, args.record_path))


print("Succesffuly finished!")




